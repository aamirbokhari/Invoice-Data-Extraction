
#wiht image processing


import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from paddleocr import PaddleOCR
import spacy
from pdf2image import convert_from_path
import cv2
import numpy as np

# pip install spacy
# python -m spacy download en_core_web_sm


PROCESSED_FILE_LIST = "processed_files1.txt"

def save_to_excel(data, file_name="invoice_data1.xlsx"):
    """
    Save the data to an Excel file and color the blank cells.
    """
    headers = ["Invoice Number", "Customer Name", "Ship To", "Date", "Balance Due", "Item", "Quantity", "Amount"]

  
    df = pd.DataFrame(data, columns=headers)
    
    if os.path.exists(file_name):
        
        existing_df = pd.read_excel(file_name)
        
        df = pd.concat([existing_df, df], ignore_index=True).drop_duplicates(subset="Invoice Number", keep="last")
    
   
    df.to_excel(file_name, index=False)
    print(f"Data saved to {file_name} successfully.")
    
   
    wb = load_workbook(file_name)
    ws = wb.active


    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
  
    for row in ws.iter_rows(min_row=2, max_col=len(headers), max_row=len(df)+1):
        for cell in row:
            if cell.value in [None, ""]: 
                cell.fill = yellow_fill
    
   
    wb.save(file_name)
    print(f"Updated {file_name} with colored blank cells.")

def image_processing(image):
    # include ur methods , which works out the best

   
    gray = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2GRAY)
    
    # # Apply thresholding to binarize the image
    # _, processed_image = cv2.threshold(gray, 128, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    
    # # Apply dilation and erosion to remove noise and small details
    # kernel = np.ones((1, 1), np.uint8)
    # processed_image = cv2.dilate(processed_image, kernel, iterations=1)
    # processed_image = cv2.erode(processed_image, kernel, iterations=1)
    
    return gray

def final_list(img):
    ocr = PaddleOCR(use_gpu=False, lang='en')
    data = ocr.ocr(img)
    
    text_list = []
    for page in data:  
        for element in page:  
            text, _ = element[1]  
            text_list.append(text)
    
    nlp = spacy.load("en_core_web_sm")
    
    def extract_entities(text_list):
        persons = []
        locations = []
        
        for item in text_list:
            doc = nlp(item)
            for ent in doc.ents:
                if ent.label_ == "PERSON" and ent.text not in persons:
                    persons.append(ent.text)
        
        concatenated_text = " ".join(text_list)
        doc = nlp(concatenated_text)
        for ent in doc.ents:
            if ent.label_ == "GPE" and ent.text not in locations:
                locations.append(ent.text)
        
        return locations

    locations = extract_entities(text_list)
    locations_string = ", ".join(locations)

    fl = []
    fl.append(text_list[2])  # Invoice Number
    
    customer_index = text_list.index('Ship Mode:')
    fl.append(text_list[customer_index+2])  # Customer Name
    
    ship_index = text_list.index(text_list[customer_index+2])
    fl.append(locations_string)  # Ship To
    
    date_index = text_list.index('Date:')
    fl.append(text_list[date_index+1])  # Date
    
    balance_index = text_list.index('Balance Due:')
    fl.append(text_list[balance_index+1])  # Balance Due
    
    item_1_index = text_list.index('Subtotal:')
    item_1_name = text_list[item_1_index-1]
    
    item_2_index = text_list.index('Amount')
    item_2_name = text_list[item_2_index+1]
    
    final_item_name = item_2_name + " " + item_1_name
    fl.append(final_item_name)  # Item
    
    qty_index = text_list.index(item_1_name)
    fl.append(text_list[qty_index-3])  # Quantity
    
    total_index = text_list.index('Total:')
    fl.append(text_list[total_index+1])  # Amount

    def extract_persons(fl):
        persons = []
        for item in fl:
            doc = nlp(item)
            for ent in doc.ents:
                if ent.label_ == "PERSON":
                    persons.append(ent.text)
        return persons

    persons = extract_persons(fl)
    if persons:
        fl[1] = persons[0]
    return fl

def load_processed_files():
    if os.path.exists(PROCESSED_FILE_LIST):
        with open(PROCESSED_FILE_LIST, "r") as file:
            return set(line.strip() for line in file)
    return set()

def save_processed_file(file_name):
    with open(PROCESSED_FILE_LIST, "a") as file:
        file.write(file_name + "\n")

if __name__ == "__main__":
    folder_path = r'C:\Users\Admin\Desktop\aamir_invoice_project\invoices'
    all_invoice_data = []

    
    processed_files = load_processed_files()

    for filename in os.listdir(folder_path):
        if filename.endswith(".pdf"):  # Ensure we only process PDF files
            if filename not in processed_files:
                file_path = os.path.join(folder_path, filename)
                
               
                images = convert_from_path(file_path)
                
                for image in images:
               
                    processed_image = image_processing(image)
                    
                    
                    invoice_data = final_list(processed_image)
                    all_invoice_data.append(invoice_data)
                
                save_processed_file(filename)
    
    save_to_excel(all_invoice_data)
